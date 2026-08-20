"""Investment / MPF position ingest.

Cash contributions that leave a bank account remain ordinary `txn` rows and can
link as transfers into an investment account. This module handles the unit
ledger: fund holdings and valuations that do not move cash and must never be
fed into `integrity.check_account`.
"""

from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

from .llm.provider import LLMProvider
from .models import Money


@dataclass
class SubaccountBalance:
    account_id: str
    member_no: str | None
    balance: Money
    allocation: Decimal | None = None


@dataclass
class FundHolding:
    instrument: str
    units: Decimal | None
    unit_price: Decimal | None
    market_value: Money
    allocation: Decimal | None = None


@dataclass
class InvestmentSnapshot:
    as_of_date: date
    scheme: str
    currency: str
    total_value: Money
    source: str
    subaccounts: list[SubaccountBalance] = field(default_factory=list)
    holdings: list[FundHolding] = field(default_factory=list)
    notes: str | None = None
    statement_file_id: str | None = None


@dataclass
class InvestmentActivity:
    account_id: str
    activity_date: date
    contribution_type: str
    activity_type: str
    amount: Money
    member_no: str | None = None
    source_hash: str = ""
    id: str = ""

    def __post_init__(self) -> None:
        if not self.id:
            basis = "|".join((
                self.account_id, self.activity_date.isoformat(),
                self.contribution_type, self.activity_type,
                str(self.amount.amount), self.amount.currency,
            ))
            self.id = hashlib.sha256(basis.encode()).hexdigest()[:32]


# Map HSBC MPF role labels -> stable account ids (must match accounts.yaml).
HSBC_MPF_ROLE_ACCOUNTS = {
    "regular employee": "hsbc_mpf_regular",
    "personal account holder": "hsbc_mpf_personal",
    "tax deductible voluntary contribution account holder": "hsbc_mpf_tdvc",
}

_MPF_ACCOUNT_ROLES = {
    "Regular Employee": "hsbc_mpf_regular",
    "Personal Account Holder": "hsbc_mpf_personal",
    "Tax Deductible Voluntary Contribution Account Holder": "hsbc_mpf_tdvc",
}
_MONEY = r"(\d[\d,]*\.\d{2})"
_DATE = r"(\d{2} [A-Z][a-z]{2} \d{4})"


def _decimal(value: str) -> Decimal:
    return Decimal(value.replace(",", ""))


def _date(value: str) -> date:
    return datetime.strptime(value, "%d %b %Y").date()


def classify_hsbc_mpf_pdf(path: str | Path) -> str:
    from .pdf.extract import extract_document
    text = extract_document(path).text
    if "MPF Member Returns" in text:
        return "member_returns"
    if "MPF Account Returns" in text:
        return "account_returns"
    if "MPF Contribution History" in text or "MPF Transaction history" in text:
        return "contribution_history"
    raise ValueError(f"{Path(path).name}: not a recognised HSBC MPF PDF")


def _parse_member_returns(path: str | Path) -> InvestmentSnapshot:
    from .pdf.extract import extract_document
    text = extract_document(path).text
    total_match = re.search(
        rf"Overall account balance.*?\n{_MONEY}\s+{_MONEY}\s+{_MONEY}", text)
    report_match = re.search(r"Balances as at (\d{1,2} [A-Z][a-z]{2} \d{4})", text)
    if not total_match or not report_match:
        raise ValueError(f"{Path(path).name}: incomplete member returns summary")
    total = _decimal(total_match.group(1))
    holdings: list[FundHolding] = []
    row = re.compile(
        rf"^(.+? Fund(?: \(with de-risking nature\))?) "
        rf"([\d,]+\.\d{{4}}) ([\d,]+\.\d{{4}}) {_MONEY}(?:\s|$)",
        re.MULTILINE,
    )
    for match in row.finditer(text):
        holdings.append(FundHolding(
            instrument=match.group(1),
            units=_decimal(match.group(2)),
            unit_price=_decimal(match.group(3)),
            market_value=Money.from_decimal(_decimal(match.group(4)), "HKD"),
        ))
    if not holdings:
        raise ValueError(f"{Path(path).name}: no MPF holdings found")
    return InvestmentSnapshot(
        as_of_date=_date(report_match.group(1)),
        scheme="hsbc_mpf",
        currency="HKD",
        total_value=Money.from_decimal(total, "HKD"),
        source="hsbc_mpf_pdf",
        holdings=holdings,
    )


def _parse_account_returns(path: str | Path) -> SubaccountBalance:
    from .pdf.extract import extract_document
    text = extract_document(path).text
    role = next((label for label in _MPF_ACCOUNT_ROLES if label in text), None)
    if role is None and "Tax Deductible Voluntary" in text:
        role = "Tax Deductible Voluntary Contribution Account Holder"
    member = re.search(r"Member account number:\s*(\d+)", text)
    summary = re.search(
        rf"Total account balance.*?\n{_MONEY}\s+{_MONEY}\s+{_MONEY}", text)
    if not role or not member or not summary:
        raise ValueError(f"{Path(path).name}: incomplete account returns")
    return SubaccountBalance(
        account_id=_MPF_ACCOUNT_ROLES[role],
        member_no=member.group(1),
        balance=Money.from_decimal(_decimal(summary.group(1)), "HKD"),
    )


def _parse_contribution_history(path: str | Path) -> list[InvestmentActivity]:
    from .pdf.extract import extract_document
    doc = extract_document(path)
    text = doc.text
    member = re.search(r"Member account number:\s*(\d+)", text)
    if not member:
        raise ValueError(f"{Path(path).name}: contribution member number missing")
    member_no = member.group(1)
    account_id = {
        "65841230": "hsbc_mpf_regular",
        "15921678": "hsbc_mpf_personal",
        "84303079": "hsbc_mpf_tdvc",
    }.get(member_no)
    if not account_id:
        raise ValueError(f"{Path(path).name}: unknown MPF member account {member_no}")
    activities: list[InvestmentActivity] = []
    pattern = re.compile(
        rf"^{_DATE}\s+(?:(.*?)\s+)?"
        rf"(Regular Contribution|Transfer In|Rebate)\s+{_MONEY}$",
        re.MULTILINE,
    )
    defaults = {
        "hsbc_mpf_personal": "Mandatory contributions from former employment(s)",
        "hsbc_mpf_tdvc": "Tax Deductible Voluntary Contributions",
    }
    for match in pattern.finditer(text):
        contribution_type = (
            (match.group(2) or "").strip()
            or defaults.get(account_id, "Contribution")
        )
        activities.append(InvestmentActivity(
            account_id=account_id,
            member_no=member_no,
            activity_date=_date(match.group(1)),
            contribution_type=contribution_type,
            activity_type=match.group(3).lower().replace(" ", "_"),
            amount=Money.from_decimal(_decimal(match.group(4)), "HKD"),
            source_hash=doc.content_hash,
        ))
    if not activities:
        raise ValueError(f"{Path(path).name}: no MPF contribution activities found")
    return activities


_MPF_LLM_SYSTEM = """You extract HSBC MPF/eMPF financial documents into JSON.
The PDF text may change layout, split rows across lines, or use revised labels.
Copy facts exactly; never estimate, calculate missing money, or invent rows.
Amounts are decimal HKD strings without currency symbols. Dates are YYYY-MM-DD.

Return one JSON object:
{
  "document_type": "member_returns" | "account_returns" | "contribution_history",
  "reported_date": "YYYY-MM-DD" | null,
  "valuation_date": "YYYY-MM-DD" | null,
  "account_role": "regular" | "personal" | "tdvc" | null,
  "member_no": "string" | null,
  "total_balance": "1234.56" | null,
  "holdings": [{
    "instrument": "exact fund name",
    "units": "decimal" | null,
    "unit_price": "decimal" | null,
    "market_value": "1234.56"
  }],
  "activities": [{
    "date": "YYYY-MM-DD",
    "contribution_type": "exact displayed contribution type",
    "activity_type": "regular_contribution" | "transfer_in" | "rebate",
    "amount": "1234.56"
  }]
}

Member Returns must contain the aggregate total and every displayed holding.
Account Returns must contain its role, member number, and total balance.
Contribution History must contain its role/member number and every displayed row.
Use null/empty arrays when a field is absent."""

_LLM_ROLE_ACCOUNTS = {
    "regular": "hsbc_mpf_regular",
    "personal": "hsbc_mpf_personal",
    "tdvc": "hsbc_mpf_tdvc",
}


def _llm_decimal(value: Any, *, field_name: str) -> Decimal:
    if value is None or isinstance(value, bool):
        raise ValueError(f"LLM MPF extraction omitted {field_name}")
    try:
        return Decimal(str(value).replace(",", ""))
    except Exception as error:
        raise ValueError(f"LLM MPF extraction returned invalid {field_name}") from error


def _llm_date(value: Any, *, field_name: str, required: bool = True) -> date | None:
    if value in (None, "") and not required:
        return None
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError) as error:
        raise ValueError(f"LLM MPF extraction returned invalid {field_name}") from error


def _parse_hsbc_mpf_pdf_llm(
    path: str | Path, provider: LLMProvider,
) -> tuple[str, InvestmentSnapshot | SubaccountBalance | list[InvestmentActivity], dict]:
    """Extract a changed MPF PDF with an LLM, then validate typed financial facts."""
    from .pdf.extract import extract_document

    document = extract_document(path)
    response = provider.complete_json(
        _MPF_LLM_SYSTEM,
        f"Filename: {Path(path).name}\n\nExtracted PDF text:\n{document.text}",
        max_tokens=6000,
    )
    data = response.data
    if not isinstance(data, dict):
        raise ValueError(f"{Path(path).name}: LLM returned no MPF document object")
    kind = str(data.get("document_type") or "")
    metadata = {
        "parser": "llm",
        "model": response.model,
        "reported_date": data.get("reported_date"),
        "valuation_date": data.get("valuation_date"),
    }
    if kind == "member_returns":
        reported = _llm_date(data.get("reported_date"), field_name="reported_date")
        valuation = _llm_date(
            data.get("valuation_date"),
            field_name="valuation_date",
            required=False,
        ) or reported
        holdings = []
        for row in data.get("holdings") or []:
            if not isinstance(row, dict) or not str(row.get("instrument") or "").strip():
                raise ValueError(f"{Path(path).name}: invalid LLM MPF holding")
            holdings.append(FundHolding(
                instrument=str(row["instrument"]).strip(),
                units=(
                    _llm_decimal(row["units"], field_name="holding units")
                    if row.get("units") is not None else None
                ),
                unit_price=(
                    _llm_decimal(row["unit_price"], field_name="holding unit price")
                    if row.get("unit_price") is not None else None
                ),
                market_value=Money.from_decimal(
                    _llm_decimal(row.get("market_value"), field_name="holding value"),
                    "HKD",
                ),
            ))
        if not holdings:
            raise ValueError(f"{Path(path).name}: LLM found no MPF holdings")
        snapshot = InvestmentSnapshot(
            as_of_date=valuation,
            scheme="hsbc_mpf",
            currency="HKD",
            total_value=Money.from_decimal(
                _llm_decimal(data.get("total_balance"), field_name="total balance"),
                "HKD",
            ),
            source="hsbc_mpf_pdf_llm",
            holdings=holdings,
            notes=(
                f"Reported {reported.isoformat()}; holdings valued "
                f"{valuation.isoformat()}; extracted by {response.model}"
            ),
        )
        return kind, snapshot, metadata
    if kind == "account_returns":
        role = str(data.get("account_role") or "").lower()
        if role not in _LLM_ROLE_ACCOUNTS:
            raise ValueError(f"{Path(path).name}: LLM returned unknown MPF account role")
        member_no = str(data.get("member_no") or "").strip()
        if not member_no:
            raise ValueError(f"{Path(path).name}: LLM omitted MPF member number")
        return kind, SubaccountBalance(
            account_id=_LLM_ROLE_ACCOUNTS[role],
            member_no=member_no,
            balance=Money.from_decimal(
                _llm_decimal(data.get("total_balance"), field_name="account balance"),
                "HKD",
            ),
        ), metadata
    if kind == "contribution_history":
        role = str(data.get("account_role") or "").lower()
        if role not in _LLM_ROLE_ACCOUNTS:
            raise ValueError(f"{Path(path).name}: LLM returned unknown MPF account role")
        member_no = str(data.get("member_no") or "").strip()
        if not member_no:
            raise ValueError(f"{Path(path).name}: LLM omitted MPF member number")
        activities = []
        for row in data.get("activities") or []:
            if not isinstance(row, dict):
                raise ValueError(f"{Path(path).name}: invalid LLM MPF activity")
            activity_type = str(row.get("activity_type") or "").lower()
            if activity_type not in {"regular_contribution", "transfer_in", "rebate"}:
                raise ValueError(f"{Path(path).name}: invalid MPF activity type")
            activities.append(InvestmentActivity(
                account_id=_LLM_ROLE_ACCOUNTS[role],
                member_no=member_no,
                activity_date=_llm_date(row.get("date"), field_name="activity date"),
                contribution_type=str(row.get("contribution_type") or "").strip(),
                activity_type=activity_type,
                amount=Money.from_decimal(
                    _llm_decimal(row.get("amount"), field_name="activity amount"),
                    "HKD",
                ),
                source_hash=document.content_hash,
            ))
        if not activities:
            raise ValueError(f"{Path(path).name}: LLM found no MPF activities")
        return kind, activities, metadata
    raise ValueError(f"{Path(path).name}: LLM returned unknown MPF document type")


def parse_hsbc_mpf_pdf_bundle(
    paths: list[str | Path],
    *, llm_provider: LLMProvider | None = None, force_llm: bool = False,
) -> tuple[InvestmentSnapshot, list[InvestmentActivity], list[dict]]:
    """Parse and cross-check MPF PDFs, with a validated LLM layout fallback."""
    parsed = []
    for raw_path in paths:
        path = Path(raw_path)
        if force_llm:
            if llm_provider is None:
                raise ValueError("LLM MPF parsing requested but no provider is configured")
            kind, value, metadata = _parse_hsbc_mpf_pdf_llm(path, llm_provider)
        else:
            try:
                kind = classify_hsbc_mpf_pdf(path)
                if kind == "member_returns":
                    value = _parse_member_returns(path)
                elif kind == "account_returns":
                    value = _parse_account_returns(path)
                else:
                    value = _parse_contribution_history(path)
                metadata = {"parser": "deterministic", "model": None}
            except ValueError:
                if llm_provider is None:
                    raise
                kind, value, metadata = _parse_hsbc_mpf_pdf_llm(path, llm_provider)
        parsed.append((path, kind, value, metadata))

    member_docs = [item for item in parsed if item[1] == "member_returns"]
    account_docs = [item for item in parsed if item[1] == "account_returns"]
    activity_docs = [item for item in parsed if item[1] == "contribution_history"]
    if len(member_docs) != 1 or len(account_docs) != 3 or not activity_docs:
        raise ValueError(
            "MPF bundle needs one Member Returns, three Account Returns, "
            "and at least one Contribution History PDF")
    snapshot = member_docs[0][2]
    if not isinstance(snapshot, InvestmentSnapshot):
        raise ValueError("MPF member returns extraction produced the wrong data type")
    subaccounts = [item[2] for item in account_docs]
    if not all(isinstance(item, SubaccountBalance) for item in subaccounts):
        raise ValueError("MPF account returns extraction produced the wrong data type")
    snapshot.subaccounts = subaccounts
    if len({item.account_id for item in snapshot.subaccounts}) != 3:
        raise ValueError("MPF bundle contains duplicate or missing member accounts")
    sub_total = sum(item.balance.amount for item in snapshot.subaccounts)
    holding_total = sum(item.market_value.amount for item in snapshot.holdings)
    if sub_total != snapshot.total_value.amount or holding_total != snapshot.total_value.amount:
        raise ValueError(
            f"MPF bundle does not reconcile: reported={snapshot.total_value.amount}, "
            f"accounts={sub_total}, holdings={holding_total}")
    # Account Returns explicitly state their holdings use 18 Aug prices; retain
    # the 19 Aug reporting date in notes but value the coherent snapshot at 18 Aug.
    if not snapshot.notes:
        snapshot.notes = (
            f"Reported {snapshot.as_of_date.isoformat()}; holdings valued "
            f"{(snapshot.as_of_date - timedelta(days=1)).isoformat()}"
        )
        snapshot.as_of_date -= timedelta(days=1)
    activities = []
    for _path, _kind, values, _metadata in activity_docs:
        if not isinstance(values, list) or not all(
            isinstance(item, InvestmentActivity) for item in values
        ):
            raise ValueError("MPF contribution extraction produced the wrong data type")
        activities.extend(values)
    docs = [
        {
            "filename": path.name,
            "classification": kind,
            "parser": metadata["parser"],
            "model": metadata["model"],
        }
        for path, kind, _value, metadata in parsed
    ]
    return snapshot, activities, docs


def parse_hsbc_mpf_position_xlsx(path: str | Path) -> InvestmentSnapshot:
    """Read the HSBC MPF position workbook captured from Personal Internet Banking."""
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(
            "MPF xlsx support needs openpyxl — install with: pip install 'finto[xlsx]'"
        ) from e

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    summary = wb["Summary"]
    funds = wb["Fund Positions"]

    total = None
    as_of = None
    currency = "HKD"
    for row in summary.iter_rows(values_only=True):
        label = (row[0] or "")
        if isinstance(label, str) and label.strip().lower() == "reported total balance":
            total = Decimal(str(row[1]))
        if isinstance(label, str) and label.strip().lower() == "position date":
            as_of = row[1]
        if isinstance(label, str) and label.strip().lower() == "currency":
            currency = str(row[1]).strip().upper()

    if total is None or as_of is None:
        raise ValueError(f"{path.name}: missing total balance or position date")
    if isinstance(as_of, datetime):
        as_of = as_of.date()
    elif isinstance(as_of, str):
        as_of = date.fromisoformat(as_of[:10])

    subaccounts: list[SubaccountBalance] = []
    for row in summary.iter_rows(values_only=True):
        role = (row[0] or "")
        if not isinstance(role, str):
            continue
        key = role.strip().lower()
        account_id = HSBC_MPF_ROLE_ACCOUNTS.get(key)
        if not account_id:
            continue
        member_no = str(row[1]).strip() if row[1] is not None else None
        bal = Money.from_decimal(Decimal(str(row[2])), currency)
        alloc = Decimal(str(row[3])) if row[3] is not None else None
        subaccounts.append(SubaccountBalance(
            account_id=account_id, member_no=member_no, balance=bal, allocation=alloc
        ))

    holdings: list[FundHolding] = []
    for row in funds.iter_rows(values_only=True):
        name = row[0]
        if not isinstance(name, str) or name.strip() in ("", "Constituent fund", "Total"):
            continue
        if "HSBC MPF" in name or name.startswith("Aggregate") or name.startswith("Reported"):
            continue
        units = Decimal(str(row[1])) if row[1] is not None else None
        price = Decimal(str(row[2])) if row[2] is not None else None
        value = Money.from_decimal(Decimal(str(row[3])), currency)
        alloc = Decimal(str(row[6])) if len(row) > 6 and row[6] is not None else None
        holdings.append(FundHolding(
            instrument=name.strip(), units=units, unit_price=price,
            market_value=value, allocation=alloc,
        ))

    return InvestmentSnapshot(
        as_of_date=as_of,
        scheme="hsbc_mpf",
        currency=currency,
        total_value=Money.from_decimal(total, currency),
        source="xlsx",
        subaccounts=subaccounts,
        holdings=holdings,
        notes=f"imported from {path.name}",
    )


def save_snapshot(conn, snap: InvestmentSnapshot, *, commit: bool = True) -> str:
    """Persist a snapshot, replacing any previous one for the same scheme+date+source."""
    snap_id = str(uuid.uuid4())
    conn.execute(
        "DELETE FROM investment_snapshot WHERE scheme=%s AND as_of_date=%s AND source=%s",
        (snap.scheme, snap.as_of_date.isoformat(), snap.source),
    )
    conn.execute(
        "INSERT INTO investment_snapshot (id, as_of_date, scheme, currency, total_value, "
        "source, statement_file_id, notes, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (snap_id, snap.as_of_date.isoformat(), snap.scheme, snap.currency,
         snap.total_value.amount, snap.source, snap.statement_file_id, snap.notes,
         datetime.now().isoformat()),
    )
    for s in snap.subaccounts:
        conn.execute(
            "INSERT INTO investment_subaccount_balance "
            "(snapshot_id, account_id, member_no, balance, currency, allocation) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (snap_id, s.account_id, s.member_no, s.balance.amount, s.balance.currency,
             str(s.allocation) if s.allocation is not None else None),
        )
    for h in snap.holdings:
        conn.execute(
            "INSERT INTO investment_holding "
            "(id, snapshot_id, instrument, units, unit_price, market_value, currency, allocation) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), snap_id, h.instrument,
             str(h.units) if h.units is not None else None,
             str(h.unit_price) if h.unit_price is not None else None,
             h.market_value.amount, h.market_value.currency,
             str(h.allocation) if h.allocation is not None else None),
        )
    if commit:
        conn.commit()
    return snap_id


def save_activities(
    conn, activities: list[InvestmentActivity], *, commit: bool = True,
) -> dict:
    inserted = 0
    skipped = 0
    for activity in activities:
        result = conn.execute(
            "INSERT INTO investment_activity "
            "(id,account_id,member_no,activity_date,contribution_type,activity_type,"
            " amount,currency,source_hash,created_at) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
            "ON CONFLICT (id) DO NOTHING",
            (
                activity.id, activity.account_id, activity.member_no,
                activity.activity_date.isoformat(), activity.contribution_type,
                activity.activity_type, activity.amount.amount,
                activity.amount.currency, activity.source_hash,
                datetime.now().isoformat(),
            ),
        )
        if result.rowcount:
            inserted += 1
        else:
            skipped += 1
    if commit:
        conn.commit()
    return {"inserted": inserted, "skipped": skipped}


def list_activities(
    conn, *, account_id: str | None = None, limit: int = 200,
) -> list[dict]:
    where = "WHERE account_id=%s" if account_id else ""
    params = (account_id, limit) if account_id else (limit,)
    rows = conn.execute(
        "SELECT id,account_id,member_no,activity_date,contribution_type,"
        "activity_type,amount,currency FROM investment_activity "
        f"{where} ORDER BY activity_date DESC,id LIMIT %s",
        params,
    )
    return [
        {
            "id": row["id"],
            "account_id": row["account_id"],
            "member_no": row["member_no"],
            "activity_date": row["activity_date"],
            "contribution_type": row["contribution_type"],
            "activity_type": row["activity_type"],
            "amount": {"amount": row["amount"], "currency": row["currency"]},
        }
        for row in rows
    ]


def list_snapshots(conn) -> list[dict]:
    rows = conn.execute(
        "SELECT id, as_of_date, scheme, currency, total_value, source, notes "
        "FROM investment_snapshot ORDER BY as_of_date DESC"
    )
    return [
        {
            "id": r["id"],
            "as_of_date": r["as_of_date"],
            "scheme": r["scheme"],
            "total": {"amount": r["total_value"], "currency": r["currency"]},
            "source": r["source"],
            "notes": r["notes"],
        }
        for r in rows
    ]


def valuation_history(conn, *, scheme: str | None = None,
                      account_id: str | None = None) -> dict:
    """Dated scheme or member-account values, ordered for charting."""
    clauses: list[str] = []
    params: list[str] = []
    if scheme:
        clauses.append("s.scheme=%s")
        params.append(scheme)
    if account_id:
        clauses.append("b.account_id=%s")
        params.append(account_id)
        value_sql = "b.balance"
        currency_sql = "b.currency"
        join_sql = (
            "JOIN investment_subaccount_balance b ON b.snapshot_id=s.id"
        )
    else:
        value_sql = "s.total_value"
        currency_sql = "s.currency"
        join_sql = ""
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = conn.execute(
        f"SELECT s.as_of_date,s.scheme,{value_sql} AS value,{currency_sql} AS currency "
        f"FROM investment_snapshot s {join_sql} {where} "
        "ORDER BY s.as_of_date ASC",
        params,
    )
    return {
        "scheme": scheme,
        "account_id": account_id,
        "points": [
            {
                "as_of_date": row["as_of_date"],
                "value": {"amount": row["value"], "currency": row["currency"]},
            }
            for row in rows
        ],
    }


def snapshot_detail(conn, snapshot_id: str) -> dict | None:
    r = conn.execute(
        "SELECT * FROM investment_snapshot WHERE id=%s", (snapshot_id,)
    ).fetchone()
    if not r:
        return None
    subs = [
        {
            "account_id": s["account_id"],
            "member_no": s["member_no"],
            "balance": {"amount": s["balance"], "currency": s["currency"]},
            "allocation": s["allocation"],
        }
        for s in conn.execute(
            "SELECT * FROM investment_subaccount_balance WHERE snapshot_id=%s",
            (snapshot_id,),
        )
    ]
    holdings = [
        {
            "instrument": h["instrument"],
            "units": h["units"],
            "unit_price": h["unit_price"],
            "market_value": {"amount": h["market_value"], "currency": h["currency"]},
            "allocation": h["allocation"],
        }
        for h in conn.execute(
            "SELECT * FROM investment_holding WHERE snapshot_id=%s ORDER BY market_value DESC",
            (snapshot_id,),
        )
    ]
    return {
        "id": r["id"],
        "as_of_date": r["as_of_date"],
        "scheme": r["scheme"],
        "total": {"amount": r["total_value"], "currency": r["currency"]},
        "source": r["source"],
        "notes": r["notes"],
        "subaccounts": subs,
        "holdings": holdings,
    }
